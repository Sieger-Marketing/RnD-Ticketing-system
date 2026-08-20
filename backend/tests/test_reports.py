"""Report generation and export (spec section 28).

The thing worth testing here is not that a file downloads -- it is that the
same report says the same thing in every format, and that "no data" survives
the trip. A KPI with no denominator returns None, and a report that renders
that as 0 would tell a manager their team achieved zero efficiency in a week
when the truth is that nobody finished anything to measure.
"""

from __future__ import annotations

import re

import io

import pytest

FORMATS = ["json", "csv", "xlsx", "pdf"]
REPORTS = ["daily", "weekly", "monthly", "breakdown"]


class TestReportAccess:
    def test_a_designer_cannot_export_reports(self, designer):
        for key in REPORTS:
            assert designer.get(f"/api/reports/{key}").status_code == 403

    def test_the_roles_that_should_export_can(self, director, manager, lead):
        for actor in (director, manager, lead):
            response = actor.get("/api/reports/daily")
            assert response.status_code == 200, actor.user["email"]

    def test_the_catalogue_describes_what_can_be_generated(self, manager):
        body = manager.get("/api/reports").json()
        assert set(body["formats"]) == set(FORMATS)
        keys = {r["key"] for r in body["reports"]}
        assert keys == set(REPORTS)
        for entry in body["reports"]:
            # The screen builds its controls from this, so each report has to
            # say which parameter it takes.
            assert entry["parameter"]
            assert entry["parameter_label"]


class TestReportFormats:
    @pytest.mark.parametrize("key", REPORTS)
    @pytest.mark.parametrize("fmt", FORMATS)
    def test_every_report_renders_in_every_format(self, manager, key, fmt):
        response = manager.get(f"/api/reports/{key}", params={"format": fmt})
        assert response.status_code == 200, response.text[:200]
        assert response.content, "empty body"

    @pytest.mark.parametrize("key", REPORTS)
    def test_downloads_are_real_files_with_a_filename(self, manager, key):
        pdf = manager.get(f"/api/reports/{key}", params={"format": "pdf"})
        assert pdf.content.startswith(b"%PDF"), "not a PDF"
        # A report that varies by parameter names the parameter in the file,
        # so several exports can sit in one downloads folder and still be
        # told apart: breakdown-product-report-20260819.pdf.
        assert re.search(
            rf'filename="{key}[a-z-]*-report-\d{{8}}\.pdf"',
            pdf.headers["content-disposition"],
        ), pdf.headers["content-disposition"]

        xlsx = manager.get(f"/api/reports/{key}", params={"format": "xlsx"})
        assert xlsx.content.startswith(b"PK"), "not a zip, so not an xlsx"

    def test_an_unknown_format_is_refused(self, manager):
        assert manager.get(
            "/api/reports/daily", params={"format": "docx"}
        ).status_code == 422


class TestNoDataSurvivesExport:
    """None means "nothing to measure", and must not become zero."""

    def test_json_keeps_null_rather_than_zero(self, manager):
        body = manager.get("/api/reports/weekly", params={"format": "json"}).json()
        section = next(
            s for s in body["sections"] if s["title"] == "Productivity by designer"
        )
        efficiency = section["columns"].index("Efficiency %")
        values = [row[efficiency] for row in section["rows"]]
        assert any(v is None for v in values), (
            "expected at least one designer with nothing to measure; if this "
            "dataset genuinely has none, the test needs a better fixture"
        )
        assert all(v is None or isinstance(v, (int, float)) for v in values)

    def test_excel_writes_a_blank_cell_not_a_zero(self, manager):
        from openpyxl import load_workbook

        content = manager.get(
            "/api/reports/weekly", params={"format": "xlsx"}
        ).content
        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook["Productivity by designer"]
        header = [c.value for c in sheet[1]]
        column = header.index("Efficiency %") + 1

        seen_blank = False
        for row in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row, column=column).value
            if value is None:
                seen_blank = True
            else:
                # Numbers must stay numeric, or the spreadsheet cannot total
                # the column the reader is looking at.
                assert isinstance(value, (int, float)), (row, value)
        assert seen_blank, "no blank cell: a missing metric was written as a value"

    def test_csv_leaves_the_field_empty(self, manager):
        import csv

        text = manager.get(
            "/api/reports/weekly", params={"format": "csv"}
        ).content.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
        start = next(
            i for i, r in enumerate(rows) if r and r[0] == "Productivity by designer"
        )
        header = rows[start + 1]
        column = header.index("Efficiency %")
        values = [r[column] for r in rows[start + 2 :] if len(r) > column]
        assert "" in values, "expected an empty cell for an unmeasurable metric"
        assert "0" not in values or "" in values


class TestReportsAgreeWithTheDashboard:
    """A report that recalculates is how a second source of truth appears."""

    def test_weekly_efficiency_matches_the_analytics_endpoint(self, manager, db):
        from datetime import date, timedelta

        from app.services import analytics_service

        today = date.today()
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)

        metrics = analytics_service.department_metrics(
            db, date_from=start, date_to=end
        )
        body = manager.get(
            "/api/reports/weekly",
            params={"format": "json", "week_of": today.isoformat()},
        ).json()
        summary = {item["label"]: item["value"] for item in body["summary"]}

        expected = metrics["efficiency_percent"]
        actual = summary["Efficiency %"]
        if expected is None:
            assert actual is None
        else:
            assert actual == pytest.approx(round(expected, 1), abs=0.05)

    def test_monthly_names_the_month_it_covers(self, manager):
        from datetime import date

        body = manager.get(
            "/api/reports/monthly",
            params={"format": "json", "month_of": "2026-03-15"},
        ).json()
        assert body["period_label"] == "March 2026"
