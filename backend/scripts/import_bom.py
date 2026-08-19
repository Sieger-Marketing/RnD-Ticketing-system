"""Import the design team's BOM release tracker.

The spreadsheet and the system agree on shape, which is why this is a mapping
rather than a translation:

    Product          -> Product
    Projects         -> Project
    Description      -> DesignRelease   (one assembly, e.g. PALLET_ASSY)
    the five phases  -> Task            (Concept, Long Lead & RM, 3D Design,
                                         Mfg. Release, Packing list & Stuffing)
    Plan / Actual    -> planned_end / actual dates
    %                -> completion_percent

Two things the sheet does not contain, and therefore neither does the import:
nobody is named against any row, and no hours are recorded anywhere. Tasks
arrive unassigned with no estimate, so capacity, utilisation and efficiency
stay empty until the team starts recording who did the work and how long it
took. That is a property of the source data, not a gap in the importer.

Dry run by default. It writes nothing without --commit.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook  # noqa: E402

# Column positions, from the merged four-row header.
COLUMNS = {
    "product": 1,
    "cars": 2,
    "project": 3,
    "dispatch": 4,
    "description": 5,
    "dsqs": 6,
    "gfc": 7,
    "input": 8,
    "concept_plan": 9,
    "concept_actual": 10,
    "longlead_plan": 11,
    "longlead_actual": 12,
    "design3d_plan": 13,
    "design3d_pct": 14,
    "design3d_actual": 15,
    "mfg_plan": 16,
    "mfg_pct": 17,
    "mfg_actual": 18,
    "packing_plan": 19,
    "packing_pct": 20,
    "packing_actual": 21,
    "remarks": 22,
}

#: The five phases, in the order the sheet lays them out. Each becomes a task
#: in the generated release, sequenced accordingly.
PHASES = [
    ("Concept", "concept_plan", "concept_actual", None, "Layout"),
    ("Long Lead & RM", "longlead_plan", "longlead_actual", None, "Coordination"),
    ("3D Design", "design3d_plan", "design3d_actual", "design3d_pct", "Modelling"),
    ("Mfg. Release", "mfg_plan", "mfg_actual", "mfg_pct", "Drawing"),
    ("Packing list & Stuffing", "packing_plan", "packing_actual", "packing_pct", "Documentation"),
]

FIRST_DATA_ROW = 5

#: Values that appear in the Projects column but do not name a project.
JUNK_PROJECT = re.compile(r"^\s*[.\-–—]?\s*$")


def as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def as_percent(value) -> float | None:
    """The sheet stores a fraction (0.95) or the word Completed."""
    if value is None or value == "":
        return None
    if isinstance(value, str):
        return 100.0 if value.strip().lower().startswith("complet") else None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return round(number * 100, 2) if number <= 1 else round(number, 2)


def is_done(value) -> bool:
    return isinstance(value, str) and value.strip().lower().startswith("complet")


@dataclass
class Row:
    product: str
    project: str
    assembly: str
    cars: int | None
    dsqs: int | None
    gfc: date | None
    dispatch: date | None
    remarks: str | None
    phases: dict = field(default_factory=dict)
    source_row: int = 0

    @property
    def project_is_usable(self) -> bool:
        """A project name that is a bare date or a full stop names nothing."""
        if not self.project or JUNK_PROJECT.match(str(self.project)):
            return False
        return as_date_like(self.project) is None


def as_date_like(value) -> date | None:
    """True when the cell holds a date rather than a name."""
    if isinstance(value, (datetime, date)):
        return as_date(value)
    return None


def read(path: Path) -> list[Row]:
    ws = load_workbook(path, data_only=True)["Sheet1"]
    rows: list[Row] = []
    product = project = None

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        cell = {k: ws.cell(row=r, column=c).value for k, c in COLUMNS.items()}
        if all(v in (None, "") for v in cell.values()):
            continue

        # Merged cells read as blank below the first row of the merge, so the
        # sheet's visual grouping is reconstructed by carrying values down.
        product = cell["product"] or product
        project = cell["project"] if cell["project"] not in (None, "") else project

        if not cell["description"]:
            continue

        phases = {}
        for name, plan_key, actual_key, pct_key, task_type in PHASES:
            plan = cell[plan_key]
            actual = cell[actual_key]
            percent = as_percent(cell[pct_key]) if pct_key else None
            if percent is None and is_done(actual):
                percent = 100.0
            if plan in (None, "") and actual in (None, "") and percent is None:
                continue
            phases[name] = {
                "task_type": task_type,
                "planned_end": as_date(plan),
                "actual_end": as_date(actual),
                "completed": is_done(actual) or as_date(actual) is not None,
                "completion_percent": percent,
            }

        rows.append(
            Row(
                product=str(product).strip() if product else "Unspecified",
                project=str(project).strip() if project else "",
                assembly=str(cell["description"]).strip(),
                cars=int(cell["cars"]) if isinstance(cell["cars"], (int, float)) else None,
                dsqs=int(cell["dsqs"]) if isinstance(cell["dsqs"], (int, float)) else None,
                gfc=as_date(cell["gfc"]),
                dispatch=as_date(cell["dispatch"]),
                remarks=str(cell["remarks"]).strip() if cell["remarks"] else None,
                phases=phases,
                source_row=r,
            )
        )
    return rows


def report(rows: list[Row]) -> None:
    print(f"Read {len(rows)} assemblies from the sheet.\n")

    products = Counter(r.product for r in rows)
    print(f"Products ({len(products)}) -> would create {len(products)} product records")
    for name, count in products.most_common():
        print(f"    {name:14} {count:4} assemblies")

    usable = [r for r in rows if r.project_is_usable]
    unusable = [r for r in rows if not r.project_is_usable]
    projects = Counter(r.project for r in usable)

    print(f"\nProjects -> {len(projects)} named, from {len(usable)} rows")
    for name, count in projects.most_common(10):
        print(f"    {name[:40]:42} {count:3} assemblies")
    if len(projects) > 10:
        print(f"    ... and {len(projects) - 10} more")

    if unusable:
        print(f"\n  {len(unusable)} rows carry no usable project name and would be SKIPPED:")
        for value, count in Counter(str(r.project) for r in unusable).most_common(6):
            print(f"    {value[:44]:46} {count:3} rows")

    print(f"\nReleases -> {len(usable)} (one per assembly on a named project)")

    phase_counts = Counter()
    for r in usable:
        for name in r.phases:
            phase_counts[name] += 1
    total_tasks = sum(phase_counts.values())
    print(f"\nTasks -> {total_tasks}, by phase:")
    for name, _, _, _, _ in PHASES:
        n = phase_counts[name]
        print(f"    {name:26} {n:4}  ({100 * n / max(len(usable), 1):5.1f}% of releases)")

    done = sum(
        1 for r in usable for p in r.phases.values() if p["completed"]
    )
    print(f"\n  {done} of {total_tasks} tasks already complete in the sheet")

    extras = {
        "Cars": sum(1 for r in usable if r.cars is not None),
        "DSQs": sum(1 for r in usable if r.dsqs is not None),
        "GFC date": sum(1 for r in usable if r.gfc),
        "Dispatch date": sum(1 for r in usable if r.dispatch),
        "Remarks": sum(1 for r in usable if r.remarks),
    }
    print("\nFields with no home in the current schema:")
    for name, count in extras.items():
        print(f"    {name:16} present on {count:4} rows")

    print("\nNot in the sheet, so not imported:")
    print("    assignee        no column names a person")
    print("    estimated hours no column records effort")
    print("    reviewer        no column records who checked the work")
    print("  Capacity, utilisation and efficiency stay empty until the team")
    print("  begins recording who did the work and how long it took.")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument(
        "--commit",
        action="store_true",
        help="Actually write. Without it this only reports what it would do.",
    )
    args = parser.parse_args()

    rows = read(args.workbook)
    report(rows)

    if not args.commit:
        print("\nDry run: nothing was written.")
        return 0

    print("\n--commit is not wired up yet: the project-name question is open.")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
