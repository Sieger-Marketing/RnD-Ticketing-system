"""Import the design team's BOM release tracker.

The spreadsheet is a wall-planner, not a table: one *block* of rows is a
project, and the block's identity is stacked vertically in a single column --
customer name on one row, the sales-order code beneath it, the work order
beneath that, then a date. Every row inside the block is one BOM item, which is
the thing the team calls a design release, and the columns to the right are the
five phases it moves through with a planned and an actual date each.

Read row by row it produces nonsense: ninety-one "projects" of which most are
order numbers. So the parser looks for where a block begins -- the Cars column
carries a value only on a block's first row -- and treats everything until the
next one as belonging to it.

Safe by default: prints what it would create and writes nothing. Pass --apply.

    python scripts/import_bom_release.py "C:/path/BOM release.xlsx"
    python scripts/import_bom_release.py "C:/path/BOM release.xlsx" --apply
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl
from sqlalchemy import select

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.enums import ReleaseStatus, TaskStatus  # noqa: E402
from app.db.session import SessionLocal  # noqa: E402
from app.models.catalog import Customer, Product, ProductFamily  # noqa: E402
from app.models.project import Project  # noqa: E402
from app.models.release import DesignRelease  # noqa: E402
from app.models.task import Task  # noqa: E402
from app.models.user import User  # noqa: E402
from app.services import code_service  # noqa: E402

FIRST_DATA_ROW = 5

# Column positions, 0-based, as the sheet lays them out.
COL_PRODUCT = 0
COL_CARS = 1
COL_PROJECT = 2
COL_DISPATCH = 3
COL_DESCRIPTION = 4
COL_DSQ = 5
COL_GFC = 6

#: The phases the tracker follows, and where each one's dates sit. These are
#: the columns the design team actually fills in, so they become the tasks --
#: importing the five standard design tasks instead would throw away every date
#: in the file, which is the only reason the file exists.
PHASES = [
    ("Concept", 8, None, 9),
    ("Long Lead & RM", 10, None, 11),
    ("3D Design", 12, 13, 14),
    ("Mfg. Release", 15, 16, 17),
    ("Packing List & Stuffing", 18, 19, 20),
]

#: A configuration heading rather than a release: "G2 0P 4C, 2150 P - HYD
#: PUZZLE - 11 UNITS". One project can hold several systems, each with its own
#: Foundation/Structures/Platform/CS beneath it, so the heading is carried onto
#: the release names -- otherwise a project ends up with four things called
#: "Structures" and no way to tell them apart. Headings never carry a DSQ
#: number, which is what makes them safe to recognise.
SYSTEM_HEADING = re.compile(r"\d+\s*[GPC].*(PUZZLE|STACKER|UNIT|CART)", re.I)

#: A Sieger sales order: AA1979, AN0288.
SO_CODE = re.compile(r"^(A[AN]\d{3,5})\b")
#: A work order in any of the several shapes the sheet uses.
WORK_ORDER = re.compile(r"(W/?O|M/WO)", re.IGNORECASE)


def _clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def _as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _as_percent(value) -> float | None:
    """The sheet stores completion as a fraction: 1 means finished, 0.95 nearly."""
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return round(number * 100, 2) if number <= 1 else round(number, 2)


def _is_name_line(line: str) -> bool:
    """Does this stacked line name a project, rather than identify an order?"""
    if not line or line in {".", "-"}:
        return False
    if SO_CODE.match(line) or WORK_ORDER.search(line):
        return False
    # A line entirely in brackets is an aside about the project above it --
    # "( Diabetic Center )" belongs to the hospital named two rows up, and
    # reading it as a project name invents one that does not exist.
    if line.startswith("("):
        return False
    return not re.fullmatch(r"[\d.\-/ :]+", line)


@dataclass(slots=True)
class Item:
    """One BOM item: a design release."""

    description: str
    sequence: int | None
    dispatch: date | None
    #: The system this belongs to, where the project holds more than one.
    system: str = ""
    phases: list[tuple[str, date | None, float | None, date | None]] = field(
        default_factory=list
    )


@dataclass(slots=True)
class Block:
    """One project, as the spreadsheet lays it out."""

    product: str = ""
    cars: int | None = None
    gfc: date | None = None
    lines: list[str] = field(default_factory=list)
    items: list[Item] = field(default_factory=list)

    @property
    def name(self) -> str:
        """The first stacked line that is not an order number or a date."""
        for line in self.lines:
            if SO_CODE.match(line) or WORK_ORDER.search(line):
                continue
            if re.fullmatch(r"[\d.\-/ ]+", line):
                continue
            if line in {".", "-"}:
                continue
            return line
        # Nothing but order numbers: use the sales order, which is at least a
        # real identifier somebody can look up.
        return self.lines[0] if self.lines else "Unnamed project"

    @property
    def sales_order(self) -> str | None:
        for line in self.lines:
            match = SO_CODE.match(line)
            if match:
                return match.group(1)
        return None

    @property
    def work_order(self) -> str | None:
        for line in self.lines:
            if WORK_ORDER.search(line):
                return line
        return None

    @property
    def customer_name(self) -> str:
        """The customer, as far as the sheet reveals one.

        Names arrive as "Ganga Medical (Coimbatore)" -- the customer and the
        site in one cell. The bracket is a location, not part of the name.
        """
        name = self.name
        name = re.sub(r"\s*\(\s*\d+\s*:\s*\d+\s*\)\s*$", "", name)  # "(01:01)"
        base = re.split(r"\s*\(", name)[0].strip()
        base = re.sub(r"^M/s\.?\s*", "", base, flags=re.IGNORECASE).strip()
        # "Ganga Medical-L cart" and "KMCH - H-cart" name the product, not the
        # customer; the customer is what is left once that is taken off.
        base = re.sub(r"\s*-\s*[A-Za-z]?\s*-?\s*cart\s*$", "", base, flags=re.IGNORECASE).strip()
        base = base.rstrip(" -")
        return base or name

    @property
    def location(self) -> str | None:
        match = re.search(r"\(([^)0-9][^)]*)\)", self.name)
        return match.group(1).strip() if match else None


def parse(path: Path) -> list[Block]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    blocks: list[Block] = []
    current: Block | None = None
    system = ""

    def start(product: str, cars, gfc) -> Block:
        nonlocal system
        system = ""
        block = Block(
            product=product,
            cars=int(cars) if str(cars).strip().isdigit() else None,
            gfc=gfc,
        )
        blocks.append(block)
        return block

    for row in sheet.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        product = _clean(row[COL_PRODUCT])
        cars = row[COL_CARS]
        project_line = _clean(row[COL_PROJECT])
        description = _clean(row[COL_DESCRIPTION])

        if not any([product, cars, project_line, description]):
            continue

        # A block begins where the car count appears -- and also where a new
        # project is simply named without one, which is how the whole tail of
        # the sheet is written. Without the second rule every H Cart, L Cart,
        # Auto Cart and Rotary project is swallowed by the last Puzzle above it.
        named = _is_name_line(project_line)
        starts_here = cars not in (None, "") or (
            named
            and current is not None
            and any(_is_name_line(line) for line in current.lines)
            and project_line != current.name
        )

        if starts_here or current is None:
            current = start(product, cars, _as_date(row[COL_GFC]))

        if project_line:
            # Some cells hold the whole identity at once rather than spreading
            # it down the rows -- separated by a line break, or by a long run
            # of spaces, because the sheet was laid out to be looked at rather
            # than read by anything. Taken as one string, the customer's name
            # ends up carrying a work order and a date.
            for part in re.split(r"\n|\s{3,}", project_line):
                part = part.strip()
                if part:
                    current.lines.append(part)
        if product and not current.product:
            current.product = product
        if current.gfc is None:
            current.gfc = _as_date(row[COL_GFC])

        if not description:
            continue

        sequence = row[COL_DSQ]
        has_sequence = str(sequence).strip().isdigit() if sequence is not None else False

        # A configuration heading names the system the rows beneath it belong
        # to; it is not itself something anyone designs.
        if not has_sequence and SYSTEM_HEADING.search(description):
            system = description
            continue

        current.items.append(
            Item(
                description=description,
                sequence=int(sequence) if has_sequence else None,
                dispatch=_as_date(row[COL_DISPATCH]),
                system=system,
                phases=[
                    (
                        label,
                        _as_date(row[plan]) if plan is not None else None,
                        _as_percent(row[pct]) if pct is not None else None,
                        _as_date(row[actual]) if actual is not None else None,
                    )
                    for label, plan, pct, actual in PHASES
                ],
            )
        )

    return [b for b in blocks if b.items]


def _status_for(percent: float | None, actual: date | None) -> str:
    if actual is not None or (percent is not None and percent >= 100):
        return TaskStatus.COMPLETED.value
    if percent:
        return TaskStatus.IN_PROGRESS.value
    return TaskStatus.NOT_STARTED.value


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="path to the BOM release spreadsheet")
    parser.add_argument("--apply", action="store_true", help="write it")
    args = parser.parse_args()

    blocks = parse(Path(args.workbook))

    customers = sorted({b.customer_name for b in blocks if b.customer_name})
    releases = sum(len(b.items) for b in blocks)
    tasks = releases * len(PHASES)

    print(f"Parsed {len(blocks)} projects from {Path(args.workbook).name}")
    print(f"  customers : {len(customers)}")
    print(f"  releases  : {releases}")
    print(f"  tasks     : {tasks}  ({len(PHASES)} phases each)")
    print()
    for block in blocks[:8]:
        print(f"  {block.name[:44]:46} {block.product:10} cars={block.cars or '-':>5}"
              f"  releases={len(block.items)}  SO={block.sales_order or '-'}")
    if len(blocks) > 8:
        print(f"  ... and {len(blocks) - 8} more")
    print()

    if not args.apply:
        print("Dry run. Nothing was written. Re-run with --apply.")
        return 0

    with SessionLocal() as db:
        creator = db.execute(select(User).order_by(User.created_at)).scalars().first()
        family = db.execute(
            select(ProductFamily).where(ProductFamily.name == "Sieger Systems")
        ).scalar_one_or_none()

        products = {p.name.lower(): p for p in db.execute(select(Product)).scalars()}
        known_customers = {
            c.name.lower(): c for c in db.execute(select(Customer)).scalars()
        }
        existing_projects = {
            p.name.lower() for p in db.execute(select(Project)).scalars()
        }

        made = {"customers": 0, "projects": 0, "releases": 0, "tasks": 0, "skipped": 0}

        for block in blocks:
            if block.name.lower() in existing_projects:
                made["skipped"] += 1
                continue

            customer = None
            if block.customer_name:
                customer = known_customers.get(block.customer_name.lower())
                if customer is None:
                    customer = Customer(
                        code=code_service.next_code(db, "customer"),
                        name=block.customer_name,
                        customer_code=code_service.next_code(db, "customer"),
                        country=block.location,
                    )
                    db.add(customer)
                    db.flush()
                    known_customers[block.customer_name.lower()] = customer
                    made["customers"] += 1

            product = products.get(block.product.lower()) if block.product else None
            if product is None and block.product and family is not None:
                product = Product(
                    code=code_service.next_code(db, "product"),
                    name=block.product,
                    product_family_id=family.id,
                )
                db.add(product)
                db.flush()
                products[block.product.lower()] = product

            project = Project(
                code=code_service.next_code(db, "project"),
                name=block.name,
                description="\n".join(block.lines),
                customer_id=customer.id if customer else None,
                product_id=product.id if product else None,
                project_type="New Design",
                priority="Medium",
                status="Design In Progress",
                car_count=block.cars,
                gfc_date=block.gfc,
                external_id=block.sales_order,
                created_by_id=creator.id if creator else None,
            )
            db.add(project)
            db.flush()
            made["projects"] += 1
            existing_projects.add(block.name.lower())

            for index, item in enumerate(block.items, start=1):
                # DSQ numbers restart at 1 for each system, so they cannot be
                # the release's position within the project.
                name = f"{item.system} - {item.description}" if item.system else item.description
                release = DesignRelease(
                    code=code_service.next_code(db, "release"),
                    project_id=project.id,
                    sequence_number=index,
                    name=name[:240],
                    release_type="Design Release",
                    product_id=product.id if product else None,
                    priority="Medium",
                    status=ReleaseStatus.PLANNING.value,
                    planned_end=item.dispatch,
                    created_by_id=creator.id if creator else None,
                )
                db.add(release)
                db.flush()
                made["releases"] += 1

                for order, (label, plan, percent, actual) in enumerate(item.phases, start=1):
                    db.add(
                        Task(
                            code=code_service.next_code(db, "task"),
                            project_id=project.id,
                            release_id=release.id,
                            sequence=order,
                            name=label,
                            task_type=label,
                            status=_status_for(percent, actual),
                            priority="Medium",
                            complexity=3,
                            planned_end=plan,
                            estimated_hours=0,
                            is_mandatory=True,
                            requires_review=False,
                            created_by_id=creator.id if creator else None,
                        )
                    )
                    made["tasks"] += 1

        db.commit()

    print("Written.")
    for key, value in made.items():
        print(f"  {key:12} {value}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
